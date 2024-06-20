import pandas as pd
from openpyxl import load_workbook
import os
import random

class ExcelRegister:
    _bookTelemetria = "Telemetria"

    def __init__(self, name_file) -> None:
        self.name_file = name_file

    def insert_data(self, frame: dict) -> None:
        df = pd.DataFrame(frame)
        if not os.path.isfile(self.name_file):
            with pd.ExcelWriter(self.name_file, engine="openpyxl") as writer:
                df.to_excel(excel_writer=writer, sheet_name=self._bookTelemetria, index=False)
        else:
            book = load_workbook(self.name_file)
            sheet = book[self._bookTelemetria]
            with pd.ExcelWriter(self.name_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                df.to_excel(excel_writer=writer, sheet_name=self._bookTelemetria, index=False, header=False, startrow=sheet.max_row)

    def _randFloat(self) -> float:
        return round(random.uniform(0, 100), 3)



def main():
    dp = ExcelRegister("Telemetria.xlsx")  # Se crea el objeto para registrar los datos
    names = ["Tiempo", "TEMP", "AX", "AY", "AZ", "GX", "GY", "GZ", "VEL", "LAT1", "LON1", "LAT2", "LON2", "PRES", "HUM", "ALT", "PILA1", "PILA2", "DIST"]
    datos = {name: [str(dp._randFloat())] for name in names}

    dp.insert_data(datos)


if __name__ == '__main__':
    main()
