import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import os
import numpy as np

from measurement import measure_time


class ExcelRegister:
    _dirName = "excel_chatgpt3"
    _bookTelemetria = "Telemetria"
    _bookEstadisticas = "Estadisticas"
    _bookGraficas = "Graficas"

    def __init__(self, name_file: str, headers: list[str]) -> None:
        self.name_file = os.path.join(self._dirName, name_file)  # Nombre del archivo de Excel.
        self.headers = headers  # Encabezados de las columnas.
        try:
            os.makedirs(self._dirName)
        except FileExistsError:
            pass
        if not os.path.isfile(self.name_file):
            with pd.ExcelWriter(self.name_file, engine="openpyxl") as writer:
                pd.DataFrame([self.headers]).to_excel(excel_writer=writer, sheet_name=self._bookTelemetria, index=False, header=False)
                pd.DataFrame().to_excel(excel_writer=writer, sheet_name=self._bookEstadisticas, index=False)
        self.columnas = ['{}1'.format(chr(ord('A') + i)) for i in range(len(self.headers))]  # Se obtienen las columnas de la hoja.
        self.pd_excel = pd.read_excel(self.name_file, sheet_name=self._bookTelemetria)

    def insert_data(self, frame: dict) -> None:
        df = pd.DataFrame(frame)
        if df.empty:
            return

        book = load_workbook(self.name_file)
        sheet = book[self._bookTelemetria]
        start_row = sheet.max_row + 1
        data = df.values.tolist()

        for i, row in enumerate(data):
            for j, value in enumerate(row):
                cell = sheet.cell(row=start_row + i, column=j + 1)
                cell.value = value

        book.save(self.name_file)
        book.close()

    def insert_stadistics(self) -> None:
        book = load_workbook(self.name_file)
        sheet = book[self._bookTelemetria]
        datos_columnas = [sheet[columna].value for columna in self.columnas if sheet[columna].value is not None]  # Se obtienen los datos de las columnas.

        # Inicia resumen de la misión
        resumen = self.pd_excel[datos_columnas].describe()
        resumen_df = pd.DataFrame(resumen)

        with pd.ExcelWriter(self.name_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            resumen_df.to_excel(excel_writer=writer, sheet_name=self._bookEstadisticas)

    def insert_graph(self, remove: str, respecto: str) -> None:
        n = self.headers.index(respecto)
        if respecto not in self.headers:
            return
        if remove == self.headers[n]:
            return
        book = load_workbook(self.name_file)
        sheet = book[self._bookTelemetria]
        datos_columnas = [sheet[columna].value for columna in self.columnas if sheet[columna].value is not None]  # Se obtienen los datos de las columnas.

        # Creando gráficas
        data_remove = datos_columnas[:]
        data_remove.remove(remove)
        filas = self.pd_excel[datos_columnas]
        new_name = f'{self._bookGraficas} ({remove})'

        if new_name not in book.sheetnames:
            book.create_sheet(title=new_name)
        sheet_removed = book[new_name]

        try:
            graph = self.one_graph(filas, remove, n)
            n -= 1
            if graph is not None:
                img = Image(graph)
                if n != 0:
                    sheet_removed.add_image(img, f"A{str(n * 25)}")
                else:
                    sheet_removed.add_image(img, f"A{str(1)}")
                book.save(self.name_file)
                book.close()
        except Exception as e:
            print(f"Error: {e}")

    def one_graph(self, filas, remove: str, n: int) -> str:
        print(f"Grafica respecto al tiempo [{n}]")
        plt.figure()  # Se crea una nueva figura.
        plt.plot(filas[remove], filas[self.headers[n]], marker="o", linestyle="-", label=f"{self.headers[n]} en funcion de {remove}")
        plt.title(f"{self.headers[n]} en función de {remove}")
        plt.xlabel(remove)
        plt.ylabel(self.headers[n])
        plt.legend()

        os.makedirs(f"Graficas ({remove})", exist_ok=True)
        graph_path = os.path.join(self._dirName, f"Graficas ({remove})", f"img{n}.png")
        plt.savefig(graph_path)
        plt.close()

        return graph_path

    def list_headers(self) -> list[str]:
        return self.headers

    @staticmethod
    def _randFloat() -> str:
        if np.random.random() < 0.05:
            return "None"
        return np.round(np.random.random() * 100, 2)


@measure_time
def main():
    names = ["Tiempo", "TEMP", "AX", "AY", "AZ", "GX", "GY", "GZ", "VEL", "LAT1", "LON1", "LAT2", "LON2", "PRES", "HUM", "ALT", "PILA1", "PILA2", "DIST"]
    dp = ExcelRegister("Telemetria.xlsx", names)  # Se crea el objeto para registrar los datos
    datos = {name: [str(ExcelRegister._randFloat())] for name in names}

    # for name in names:
    #     dp.insert_graph("Tiempo", name)
    # dp.insert_data(datos)
    # dp.insert_stadistics()


if __name__ == '__main__':
    main()
